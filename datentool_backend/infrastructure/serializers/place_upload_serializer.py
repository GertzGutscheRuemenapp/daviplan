from io import StringIO
import os
from collections import OrderedDict
from typing import Tuple
import pandas as pd


from django.conf import settings
from django.db.models.fields import FloatField
from django.contrib.gis.geos import Point
from django.contrib.gis.db.models.functions import GeoFunc, Transform
from rest_framework import serializers

from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl import styles
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.dimensions import ColumnDimension, RowDimension
from openpyxl.worksheet.datavalidation import DataValidation

from datentool_backend.user.models import (Infrastructure,

                                           )

from datentool_backend.area.models import FClass, FieldType, FieldTypes
from datentool_backend.infrastructure.models import (PlaceAttribute,
                                                     Place,
                                                     PlaceField,
                                                     Capacity,
                                                     )



infrastructure_id_serializer = serializers.PrimaryKeyRelatedField(
    queryset=Infrastructure.objects.all(),
    help_text='infrastructure_id',)


class St_X(GeoFunc):
    function='ST_X'
    output_field = FloatField()


class St_Y(GeoFunc):
    function='ST_Y'
    output_field = FloatField()


class PlacesTemplateSerializer(serializers.Serializer):
    """Serializer for uploading Places for an Infrastructure"""
    excel_file = serializers.FileField()
    drop_constraints = serializers.BooleanField(default=True)

    def create_template(self, infrastructure_id: int) -> bytes:
        """Create a template file for places of the give infrastructure"""

        infrastructure = Infrastructure.objects.get(pk=infrastructure_id)

        excel_filename = 'Standorte_und_Kapazitäten.xlsx'

        sn_classifications = 'Klassifizierungen'

        columns = {'Name': 'So werden die Einrichtungen auf den Karten beschriftet. '\
                   'Jeder Standort muss einen Namen haben, den kein anderer Standort trägt.',
                  'Straße': 'Postalisch korrekte Schreibweise',
                  'Hausnummer': 'Zahl, ggf. mit Buchstaben (7b) oder 11-15',
                  'PLZ': 'fünfstellig',
                  'Ort': 'Postalisch korrekte Schreibweise',
                  'Lon': 'Längengrad, in WGS84',
                  'Lat': 'Breitengrad, in WGS84',}

        dv_01 = DataValidation(type="whole",
                        operator="between",
                        formula1=0,
                        formula2=1,
                        allow_blank=True)
        dv_01.error = 'Nur 0 oder 1 erlaubt'
        dv_01.title = 'Ungültige 0/1-Werte'

        dv_float = DataValidation(type="decimal",
                                allow_blank=True)
        dv_float.error = 'Nur Zahlen erlaubt'
        dv_float.title = 'Ungültige Zahlen-Werte'

        dv_pos_float = DataValidation(type="decimal",
                                operator="greaterThanOrEqual",
                                formula1=0,
                                allow_blank=True)
        dv_pos_float.error = 'Nur Zahlen >= 0 erlaubt'
        dv_pos_float.title = 'Ungültige positive Zahlen-Werte'

        validations = {}
        classifications = OrderedDict()
        col_no_classifications = 2

        df_places = pd.DataFrame(columns=pd.Index(columns.items()),
                                 index=pd.Index([], dtype='int64', name='place_id'))
        df_places.index.name = 'place_id'
        place_attrs = PlaceAttribute\
            .value_annotated_qs()\
            .filter(place__infrastructure=infrastructure)
        labels = pd.DataFrame(place_attrs\
            .filter(field__name=infrastructure.label_field)\
            .values('place_id', '_value'),
            columns=['place_id', '_value'])\
            .set_index('place_id')
        df_places['Name'] = labels['_value']

        for col in ['Straße', 'Hausnummer', 'PLZ', 'Ort']:
            try:
                infrastructure.placefield_set.get(name=col)
                attrs = pd.DataFrame(place_attrs\
                    .filter(field__name=col)\
                    .values('place_id', '_value'),
                    columns=['place_id', '_value'])\
                    .set_index('place_id')
                df_places[col] = attrs['_value']

            except PlaceField.DoesNotExist:
                continue

        place_coords = Place.objects.filter(infrastructure=infrastructure)\
            .annotate(pnt_wgs84=Transform('geom', 4326))\
            .annotate(Lon=St_X('pnt_wgs84'), Lat=St_Y('pnt_wgs84'))

        df_coords = pd.DataFrame(place_coords\
                                 .values('id', 'Lon', 'Lat'),
                                 columns=['id', 'Lon', 'Lat'])\
                                 .set_index('id')

        df_places[['Lon', 'Lat']] = df_coords

        col_no = len(df_places.columns) + 1

        for place_field in infrastructure.placefield_set.all():
            col = place_field.name
            if col in columns or col == infrastructure.label_field:
                continue
            col_no += 1
            if place_field.field_type.ftype == FieldTypes.STRING:
                description = f'Nutzerdefinierte Spalte (Text)'
                value_col = '_value'

            elif place_field.field_type.ftype == FieldTypes.NUMBER:
                description = f'Nutzerdefinierte Spalte (Zahl)'
                validations[col_no] = dv_float
                value_col = 'num_value'

            else:
                description = f'Nutzerdefinierte Spalte (Klassifizierte Werte)'
                cf_colno = classifications.get(place_field.field_type.name)
                value_col = '_value'

                if cf_colno:
                    df, cn = cf_colno
                else:
                    df = pd.DataFrame(
                        place_field.field_type.fclass_set.values('order', 'value'))\
                        .set_index('order')\
                        .rename(columns={'value': place_field.field_type.name,})

                    cn = col_no_classifications
                    classifications[place_field.field_type.name] = (df, cn)
                    col_no_classifications += 1

                # list validator
                letter = get_column_letter(cn)
                dv_range = f'={quote_sheetname(sn_classifications)}!${letter}$2:${letter}$9999'
                dv = DataValidation(type='list', formula1=dv_range, allow_blank=True)
                dv.error = 'Nur Werte aus Liste erlaubt'
                dv.title = f'Liste für {place_field.field_type.name}'
                validations[col_no] = dv

            attrs = pd.DataFrame(place_attrs\
                .filter(field__name=col)\
                .values('place_id', value_col),
                columns=['place_id', value_col])\
                .set_index('place_id')
            df_places.loc[:, (col, description)] = attrs[value_col]

        for service in infrastructure.service_set.all():
            col_no += 1
            if service.has_capacity:
                col = f'Kapazität für Leistung {service.name}'
                description = service.capacity_plural_unit
                dv = dv_pos_float
            else:
                col = f'Bietet Leistung {service.name} an'
                description = '1 wenn ja, sonst 0'
                dv = dv_01

            capacities = pd.DataFrame(Capacity.objects\
                .filter(service=service, from_year=0, scenario=None)\
                .values('place_id', 'capacity'),
                columns=['place_id', 'capacity'])\
                .set_index('place_id')
            df_places.loc[:, (col, description)] = capacities['capacity']
            validations[col_no] = dv

        if classifications:
            df_classification = pd.concat(
            [c[0] for c in classifications.values()],
            axis=1)
        else:
            df_classification = pd.DataFrame(index=pd.Index([], name='order'))


        fn = os.path.join(settings.MEDIA_ROOT, excel_filename)
        sheetname = 'Standorte und Kapazitäten'
        with pd.ExcelWriter(fn, engine='openpyxl') as writer:
            df_meta = pd.DataFrame({'value': [infrastructure_id]},
                                   index=pd.Index(['infrastructure'], name='key'))
            df_meta.to_excel(writer, sheet_name='meta')

            df_classification.reset_index().to_excel(writer,
                                                     sheet_name=sn_classifications,
                                                     index=False)

            df_places.to_excel(writer,
                               sheet_name=sheetname,
                               freeze_panes=(3, 2))

            ws: Worksheet = writer.sheets.get(sheetname)

            row12 = ws[1:2]
            for row in row12:
                for cell in row:
                    cell.alignment = styles.Alignment(horizontal='center', wrap_text=True)

            dv = DataValidation(type="decimal",
                                operator="between",
                                formula1=0,
                                formula2=90,
                                allow_blank=True)

            dv.error ='Koordinaten müssen in WGS84 angegeben werden und zwischen 0 und 90 liegen'
            dv.errorTitle = 'Ungültige Koordinaten'
            dv.add('G3:H999999')
            ws.add_data_validation(dv)
            ws.add_data_validation(dv_01)
            ws.add_data_validation(dv_float)
            ws.add_data_validation(dv_pos_float)
            for col_no, dv in validations.items():
                letter = get_column_letter(col_no)
                cell_range = f'{letter}3:{letter}999999'
                dv.add(cell_range)
                if not dv in ws.data_validations:
                    ws.add_data_validation(dv)

            ws.row_dimensions[3] = RowDimension(ws, index=3, hidden=True)

            ws.column_dimensions['A'] = ColumnDimension(ws, index='A', hidden=True)
            ws.column_dimensions['B'] = ColumnDimension(ws, index='B', width=30)
            for col_no in range(3, len(df_places.columns) + 2):
                col = get_column_letter(col_no)
                ws.column_dimensions[col] = ColumnDimension(ws, index=col, width=16)

        content = open(fn, 'rb').read()
        return content

    def read_excel_file(self, request, infrastructure_id: int) -> pd.DataFrame:
        """read excelfile and return a dataframe"""
        excel_file = request.FILES['excel_file']

        # check if the right Excel-File is uploaded
        df_meta = pd.read_excel(excel_file.file,
                           sheet_name='meta').set_index('key')

        infra_id = df_meta.loc['infrastructure', 'value']
        try:
            infra = Infrastructure.objects.get(pk=infra_id)
        except Infrastructure.DoesNotExist:
            raise ValueError(f'Excel-File is for infrastructure with id {infra_id}, '
                             f'which does not exist')
        msg = f'file is for infrastructure {infra.name} with id {infra_id}, but request was sent to id {infrastructure_id}'
        assert int(infrastructure_id) == df_meta.loc['infrastructure', 'value'], msg

        # Check the classification

        df_klassifizierungen = pd.read_excel(excel_file.file,
                           sheet_name='Klassifizierungen').set_index('order')

        for field_name, series in df_klassifizierungen.items():
            ft, created = FieldType.objects.get_or_create(name=field_name,
                                                 ftype=FieldTypes.CLASSIFICATION)
            series = series.loc[~pd.isna(series)]
            # add/change entries of the classification values
            for order, value in series.iteritems():
                fc, created = FClass.objects.get_or_create(ftype=ft, order=order)
                fc.value = value
                fc.save()

        # get the services and the place fields of the infrastructure
        services = infra.service_set.all().values('id', 'has_capacity', 'name')
        place_fields = infra.placefield_set.all().values('id',
                                                         'name',
                                                         'field_type',
                                                         'field_type__ftype',
                                                         #'field_type__fclass',
                                                         )


        # get a mapping of the classifications
        fclass_values = FClass.objects\
            .filter(ftype__in=place_fields.values_list('field_type'))\
            .values('ftype', 'id', 'value')

        fclass_dict = {(fclass_value['ftype'], fclass_value['value']):
                       fclass_value['id']
                       for fclass_value in fclass_values}


        # collect the places, attributes and capacities in lists
        place_ids = []
        place_attributes = []
        capacities = []

        # read the excel-file

        df_places = pd.read_excel(excel_file.file,
                           sheet_name='Standorte und Kapazitäten',
                           skiprows=[1, 2]).set_index('Unnamed: 0')
        df_places.index.name = 'place_id'

        # iterate over all places
        for place_id, place_row in df_places.iterrows():
            if pd.isna(place_id):
                place_id = None
            # check if place_id exists, otherwise create it
            try:
                place = Place.objects.get(pk=place_id)
            except Place.DoesNotExist:
                place = Place()

            place_name = place_row['Name']
            if pd.isna(place_name):
                place_name = ''

            #  do a geocoding, if no coordinates are provided
            lon = place_row['Lon']
            lat = place_row['Lat']
            if pd.isna(lon) or pd.isna(lat):
                lon, lat = self.geocode(place_row)

            # create the geometry and transform to WebMercator
            geom = Point(lon, lat, srid=4326)
            geom.transform(3857)

            # set the place columns
            place.infrastructure = infra
            place.name = place_name
            place.geom=geom
            place.save()
            place_ids.append(place.id)

            # collect the place_attributes
            for place_field in place_fields:
                place_field_name = place_field['name']
                attr = place_row.get(place_field_name)
                if attr and not pd.isna(attr):
                    str_value, num_value, class_value = None, None, None
                    ftype = place_field['field_type__ftype']
                    if ftype == 'STR':
                        str_value = attr
                    elif ftype == 'NUM':
                        num_value = float(attr)
                    else:
                        class_value = fclass_dict[(place_field['field_type'], attr)]
                    attribute = (place.id, place_field['id'], str_value, num_value, class_value)
                    place_attributes.append(attribute)

            # collect the capacities
            for service in services:
                service_name = service['name']
                if service['has_capacity']:
                    col = f'Kapazität für Leistung {service_name}'
                else:
                    col = f'Bietet Leistung {service_name} an'
                capacity = place_row.get(col)
                if capacity is not None and not pd.isna(capacity):
                    capacities.append((place.id, service['id'], capacity, 0, 99999999))


        # upload the place-attributes
        df_place_attributes = pd.DataFrame(place_attributes, columns=[
            'place_id', 'field_id', 'str_value', 'num_value', 'class_value_id'
        ])

        df_place_attributes['class_value_id'] = df_place_attributes['class_value_id'].astype('Int64')

        existing_place_attrs = PlaceAttribute.objects.filter(place_id__in=place_ids)
        existing_place_attrs.delete()

        if len(df_place_attributes):
            with StringIO() as file:
                df_place_attributes.to_csv(file, index=False)
                file.seek(0)
                PlaceAttribute.copymanager.from_csv(
                    file,
                    drop_constraints=False, drop_indexes=False,
                )


        # upload the capacities
        df_capacities = pd.DataFrame(capacities,
                                     columns=['place_id', 'service_id', 'capacity',
                                              'from_year', 'to_year'])

        existing_capacities = Capacity.objects.filter(place_id__in=place_ids)
        existing_capacities.delete()

        if len(df_capacities):
            with StringIO() as file:
                df_capacities.to_csv(file, index=False)
                file.seek(0)
                Capacity.copymanager.from_csv(
                    file,
                    drop_constraints=False, drop_indexes=False,
                )

        df = pd.DataFrame()
        return df

    def geocode(self, place_row: dict) -> Tuple[float, float]:
        """do the geocoding, if no coordinates are provided"""
        raise NotImplementedError('Geocoding is not implemented yet. '
                                  'Please provide Lon/Lat-coordinates')
from io import StringIO
import os
from collections import OrderedDict
from typing import Tuple
import pandas as pd
import logging

from django.conf import settings
from django.db.models.fields import FloatField
from django.core.exceptions import BadRequest
from django.contrib.gis.geos import Point
from django.contrib.gis.db.models.functions import GeoFunc, Transform
from rest_framework import serializers, status
from rest_framework.response import Response

from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl import styles
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.worksheet.datavalidation import DataValidation

from datentool_backend.site.models import SiteSetting
from datentool_backend.area.models import FClass, FieldTypes
from datentool_backend.infrastructure.models.places import (PlaceAttribute,
                                                            Place,
                                                            PlaceField,
                                                            Capacity)
from datentool_backend.infrastructure.models.infrastructures import (
    Infrastructure)
from datentool_backend.utils.crypto import decrypt
from datentool_backend.utils.bkg_geocoder import BKGGeocoder
from datentool_backend.utils.processes import (ProtectedProcessManager,
                                               ProcessScope)

infrastructure_id_serializer = serializers.PrimaryKeyRelatedField(
    queryset=Infrastructure.objects.all(),
    help_text='infrastructure_id',)


class St_X(GeoFunc):
    function='ST_X'
    output_field = FloatField()


class St_Y(GeoFunc):
    function='ST_Y'
    output_field = FloatField()

ADDRESS_FIELDS = {
    'Straße': 'Postalisch korrekte Schreibweise',
    'Hausnummer': 'Zahl, ggf. mit Buchstaben (7b) oder 11-15',
    'PLZ': 'fünfstellig',
    'Ort': 'Postalisch korrekte Schreibweise'
}


class PlacesTemplateSerializer(serializers.Serializer):
    """Serializer for uploading Places for an Infrastructure"""
    excel_file = serializers.FileField()
    drop_constraints = serializers.BooleanField(default=False)
    logger = logging.getLogger('infrastructure')
    scope = ProcessScope.INFRASTRUCTURE

    def create_template(self, infrastructure_id: int) -> bytes:
        """Create a template file for places of the give infrastructure"""

        infrastructure = Infrastructure.objects.get(pk=infrastructure_id)

        excel_filename = 'Standorte_und_Kapazitäten.xlsx'

        sn_classifications = 'Klassifizierungen'

        columns = {'Name': 'So werden die Einrichtungen auf den Karten beschriftet. '\
                   'Jeder Standort muss einen Namen haben, den kein anderer Standort trägt.'}

        columns.update(ADDRESS_FIELDS)
        columns.update({'Lon': 'Längengrad, in WGS84',
                        'Lat': 'Breitengrad, in WGS84'})

        dv_01 = DataValidation(type="whole",
                        operator="between",
                        formula1=0,
                        formula2=1,
                        allow_blank=True)
        dv_01.error = 'Nur 0 oder 1 erlaubt'
        dv_01.title = 'Ungültige 0/1-Werte'

        dv_float = DataValidation(type="decimal",
                                allow_blank=True)
        dv_float.error = 'Nur Zahlen erlaubt'
        dv_float.title = 'Ungültige Zahlen-Werte'

        dv_pos_float = DataValidation(type="decimal",
                                operator="greaterThanOrEqual",
                                formula1=0,
                                allow_blank=True)
        dv_pos_float.error = 'Nur Zahlen >= 0 erlaubt'
        dv_pos_float.title = 'Ungültige positive Zahlen-Werte'

        validations = {}
        classifications = OrderedDict()
        col_no_classifications = 2

        df_places = pd.DataFrame(columns=pd.Index(columns.items()),
                                 index=pd.Index([], dtype='int64', name='place_id'))
        place_id_description = 'Hier stehen die daviplan-internen Nummern von bereits '\
            'hochgeladenen Standorten. Bitte nicht verändern.'
        df_places.index.name = ('place_id', place_id_description)
        place_attrs = PlaceAttribute\
            .value_annotated_qs()\
            .filter(place__infrastructure=infrastructure)

        places = Place.objects.filter(infrastructure=infrastructure)
        if places:
            df_placename = pd.DataFrame(places.values('id', 'name'))\
                .rename(columns={'id': 'place_id', })\
                .set_index('place_id')
            df_places['Name'] = df_placename['name']
        for col in ADDRESS_FIELDS.keys():
            try:
                infrastructure.placefield_set.get(name=col)
                attrs = pd.DataFrame(place_attrs
                    .filter(field__name=col)
                    .values('place_id', '_value'),
                    columns=['place_id', '_value'])\
                    .set_index('place_id')
                df_places[col] = attrs['_value']

            except PlaceField.DoesNotExist:
                continue

        place_coords = Place.objects.filter(infrastructure=infrastructure)\
            .annotate(pnt_wgs84=Transform('geom', 4326))\
            .annotate(Lon=St_X('pnt_wgs84'), Lat=St_Y('pnt_wgs84'))

        df_coords = pd.DataFrame(place_coords\
                                 .values('id', 'Lon', 'Lat'),
                                 columns=['id', 'Lon', 'Lat'])\
                                 .set_index('id')

        df_places[['Lon', 'Lat']] = df_coords

        col_no = len(df_places.columns) + 1

        for place_field in infrastructure.placefield_set.all():
            col = place_field.name
            if col in columns or col == infrastructure.label_field:
                continue
            col_no += 1
            if place_field.field_type.ftype == FieldTypes.STRING:
                description = f'Nutzerdefinierte Spalte (Text)'
                value_col = '_value'

            elif place_field.field_type.ftype == FieldTypes.NUMBER:
                unit = place_field.unit or 'Zahl'
                description = f'Nutzerdefinierte Spalte ({unit})'
                validations[col_no] = dv_float
                value_col = 'num_value'

            else:
                description = f'Nutzerdefinierte Spalte (Klassifizierte Werte)'

                cl_name = place_field.field_type.name
                cf_colno = classifications.get(cl_name)
                value_col = '_value'

                if cf_colno:
                    df, cn = cf_colno
                else:
                    df = pd.DataFrame(
                        place_field.field_type.fclass_set.values('order', 'value'))\
                        .set_index('order')\
                        .rename(columns={'value': cl_name,})

                    cn = col_no_classifications
                    classifications[cl_name] = (df, cn)
                    col_no_classifications += 1

                # list validator
                letter = get_column_letter(cn)
                dv_range = f'={quote_sheetname(sn_classifications)}!${letter}$2:${letter}$9999'
                dv = DataValidation(type='list', formula1=dv_range, allow_blank=True)
                dv.error = 'Nur Werte aus Liste erlaubt'
                dv.title = f'Liste für {place_field.field_type.name}'
                validations[col_no] = dv

            attrs = pd.DataFrame(place_attrs\
                .filter(field__name=col)\
                .values('place_id', value_col),
                columns=['place_id', value_col])\
                .set_index('place_id')
            df_places.loc[:, (col, description)] = attrs[value_col]

        for service in infrastructure.service_set.all():
            col_no += 1
            if service.has_capacity:
                col = f'Kapazität für Leistung {service.name}'
                description = service.capacity_plural_unit
                dv = dv_pos_float
            else:
                col = f'Bietet Leistung {service.name} an'
                description = '1 wenn ja, sonst 0'
                dv = dv_01

            capacities = pd.DataFrame(Capacity.objects\
                .filter(service=service, from_year=0, scenario=None)\
                .values('place_id', 'capacity'),
                columns=['place_id', 'capacity'])\
                .set_index('place_id')
            df_places.loc[:, (col, description)] = capacities['capacity']
            validations[col_no] = dv

        if classifications:
            df_classification = pd.concat(
            [c[0] for c in classifications.values()],
            axis=1)
        else:
            df_classification = pd.DataFrame(index=pd.Index([], name='order'))


        fn = os.path.join(settings.MEDIA_ROOT, excel_filename)
        sheetname = 'Standorte und Kapazitäten'
        with pd.ExcelWriter(fn, engine='openpyxl') as writer:
            meta = writer.book.create_sheet('meta')
            meta['A1'] = 'infrastructure'
            meta['B1'] = infrastructure.name
            meta.sheet_state = 'hidden'

            df_classification.reset_index().to_excel(writer,
                                                     sheet_name=sn_classifications,
                                                     index=False)
            # hide classifications
            sheet = writer.book.get_sheet_by_name(sn_classifications)
            sheet.sheet_state = 'hidden'

            df_places\
                .reset_index()\
                .to_excel(writer,
                          sheet_name=sheetname,
                          freeze_panes=(2, 2))

            ws: Worksheet = writer.sheets.get(sheetname)

            ws.delete_rows(3, 1)
            ws.delete_cols(1, 1)

            grey_fill = styles.PatternFill("solid", fgColor='00C0C0C0')
            ws.column_dimensions['A'].fill = grey_fill
            col_placeid = ws['A1:A99999']
            for row in col_placeid:
                row[0].fill = grey_fill

            ws.column_dimensions['E'].number_format = '@'
            col_plz = ws['E1:E99999']
            for row in col_plz:
                row[0].number_format = '@'

            row12 = ws[1:2]
            for row in row12:
                for cell in row:
                    cell.alignment = styles.Alignment(horizontal='center', wrap_text=True)

            dv = DataValidation(type="decimal",
                                operator="between",
                                formula1=0,
                                formula2=90,
                                allow_blank=True)

            dv.error ='Koordinaten müssen in WGS84 angegeben werden und zwischen 0 und 90 liegen'
            dv.errorTitle = 'Ungültige Koordinaten'
            dv.add('G3:H999999')

            dv_unique_name = DataValidation(type='custom',
                                            formula1='=COUNTIF(B$3:B$99999,B3)<2',
                                            allow_blank=True)
            dv_unique_name.error ='Name muss eindeutig sein'
            dv_unique_name.errorTitle = 'Name nicht eindeutig'
            dv_unique_name.add('B3:B999999')

            dv_plz = DataValidation(type='textLength',
                                    operator='equal',
                                    formula1=5,
                                    allow_blank=True)
            dv_plz.error ='Postleitzahl muss fünfstellig sein'
            dv_plz.errorTitle = 'Postleitzahl_Format'
            dv_plz.add('E3:E999999')

            ws.add_data_validation(dv)
            ws.add_data_validation(dv_unique_name)
            ws.add_data_validation(dv_01)
            ws.add_data_validation(dv_float)
            ws.add_data_validation(dv_pos_float)
            ws.add_data_validation(dv_plz)
            for col_no, dv in validations.items():
                letter = get_column_letter(col_no)
                cell_range = f'{letter}3:{letter}999999'
                dv.add(cell_range)
                if not dv in ws.data_validations:
                    ws.add_data_validation(dv)

            ws.column_dimensions['A'] = ColumnDimension(ws, index='A', width=20)
            ws.column_dimensions['B'] = ColumnDimension(ws, index='B', width=30)
            for col_no in range(3, len(df_places.columns) + 2):
                col = get_column_letter(col_no)
                ws.column_dimensions[col] = ColumnDimension(ws, index=col, width=16)

        content = open(fn, 'rb').read()
        return content

    def read_excel_file(self, request) -> pd.DataFrame:
        """read excelfile and return a dataframe"""
        excel_file = request.FILES['excel_file']
        infrastructure_id = request.data.get('infrastructure')
        run_sync = bool(strtobool(request.data.get('sync', 'False')))

        if infrastructure_id is None:
            raise Exception(f'Die ID der Infrastruktur muss im Formular mit übergeben werden.')

        with ProtectedProcessManager(
                user=request.user,
                scope=getattr(serializer, 'scope', ProcessScope.INFRASTRUCTURE)) as ppm:
            try:
                ppm.run(self.process_place_upload,
                        infrastructure_id, excel_file,
                        self.logger, sync=run_sync)
            except Exception as e:
                return Response({'message': str(e)},
                                status.HTTP_500_INTERNAL_SERVER_ERROR)

        return Response({'message': 'Upload/Geocoding gestartet'},
                        status=status.HTTP_202_ACCEPTED)


    @staticmethod
    def process_place_upload(infrastructure_id: int, excel_file: str, logger):
        ProcessPlaceUpload(logger).process_place_upload(infrastructure_id, excel_file)


class ProcessPlaceUpload:

    def __init__(self, logger):
        self.logger = logger

    def process_place_upload(self, infrastructure_id: int, excel_file: str):
        infra = Infrastructure.objects.get(pk=infrastructure_id)

        # get the services and the place fields of the infrastructure
        services = infra.service_set.all().values('id', 'has_capacity', 'name')
        place_fields = infra.placefield_set.all().values('id',
                                                         'name',
                                                         'field_type',
                                                         'field_type__ftype',
                                                         'field_type__name',
                                                         )

        # get a mapping of the classifications
        fclass_values = FClass.objects\
            .filter(ftype__in=place_fields.values_list('field_type'))\
            .values('ftype', 'id', 'value')

        fclass_dict = {(fclass_value['ftype'], fclass_value['value']):
                       fclass_value['id']
                       for fclass_value in fclass_values}

        # collect the places, attributes and capacities in lists
        place_ids = []
        place_attributes = []
        capacities = []

        # read the excel-file
        df_places = pd.read_excel(excel_file.file,
                           sheet_name='Standorte und Kapazitäten',
                           skiprows=[1])\
            .set_index('place_id')

        # delete places that have no place_id in the excel-file
        place_ids_in_excelfile = df_places.index[~pd.isna(df_places.index)]
        places_to_delete = Place.objects.filter(infrastructure=infrastructure_id)\
            .exclude(id__in=place_ids_in_excelfile)
        n_elems_deleted, elems_deleted = places_to_delete.delete()
        n_places_deleted = elems_deleted.get('datentool_backend.Place')
        self.logger.info(f'{n_places_deleted or 0} bestehende Standorte gelöscht')

        # get BKG-Geocoding-Key
        site_settings = SiteSetting.load()
        UUID = site_settings.bkg_password or None
        bkg_error = ('Es wurde keine UUID für die Nutzung des BKG-'
                     'Geokodierungsdienstes hinterlegt' if not UUID else '')
        geocoder = None
        if UUID:
            try:
                bkg_pass = decrypt(UUID)
            except Exception:
                bkg_error = ('Die UUID des BKG-Geokodierungsdienstes konnte '
                             'nicht entschlüsselt werden. Vermutlich ist der '
                             'hinterlegte Schlüssel (ENCRYPT_KEY) nicht valide '
                             '(URL-safe base64-encoded 32-byte benötigt)')
            else:
                geocoder = BKGGeocoder(bkg_pass, crs='EPSG:3857')

        # iterate over all places
        n_new = 0
        for place_id, place_row in df_places.iterrows():
            if pd.isna(place_id):
                place_id = None
            # check if place_id exists, otherwise create it
            try:
                place = Place.objects.get(pk=place_id)
            except Place.DoesNotExist:
                place = Place()
                n_new += 1

            place_name = place_row['Name']
            if pd.isna(place_name):
                place_name = ''

            #  do a geocoding, if no coordinates are provided
            lon = place_row['Lon']
            lat = place_row['Lat']
            if pd.isna(lon) or pd.isna(lat):
                if geocoder:
                    # ToDo: handle auth errors ...
                    self.logger.info('Geokodiere Adressen')
                    res = self.geocode(geocoder, place_row)
                    if res:
                        lon, lat = res
                        geom = Point(lon, lat, srid=3857)
                    else:
                        # ToDo: ????
                        pass
                else:
                    raise ConnectionError(bkg_error)
            else:
                # create the geometry and transform to WebMercator
                geom = Point(lon, lat, srid=4326)
                geom.transform(3857)

            # set the place columns
            place.infrastructure = infra
            place.name = place_name
            place.geom = geom
            place.save()
            place_ids.append(place.id)

            # collect the place_attributes
            for place_field in place_fields:
                place_field_name = place_field['name']
                attr = place_row.get(place_field_name)
                if attr and not pd.isna(attr):
                    str_value, num_value, class_value = None, None, None
                    ftype = place_field['field_type__ftype']
                    if ftype == 'STR':
                        str_value = attr
                    elif ftype == 'NUM':
                        num_value = float(attr)
                    else:
                        try:
                            class_value = fclass_dict[(place_field['field_type'],
                                                       attr)]
                        except KeyError:
                            fieldtype_name = place_field['field_type__name']
                            msg = f'Wert {attr} existiert nicht für Klassifizierung '\
                                f'{fieldtype_name} in Spalte {place_field_name}'
                            raise BadRequest(msg)
                    attribute = (place.id, place_field['id'], str_value, num_value, class_value)
                    place_attributes.append(attribute)

            # collect the capacities
            for service in services:
                service_name = service['name']
                if service['has_capacity']:
                    col = f'Kapazität für Leistung {service_name}'
                else:
                    col = f'Bietet Leistung {service_name} an'
                capacity = place_row.get(col)
                if capacity is not None and not pd.isna(capacity):
                    capacities.append((place.id, service['id'], capacity, 0, 99999999))

        # upload the place-attributes
        df_place_attributes = pd.DataFrame(place_attributes, columns=[
            'place_id', 'field_id', 'str_value', 'num_value', 'class_value_id'
        ])

        df_place_attributes['class_value_id'] = df_place_attributes['class_value_id'].astype('Int64')

        existing_place_attrs = PlaceAttribute.objects.filter(place_id__in=place_ids)
        existing_place_attrs.delete()

        if len(df_place_attributes):
            with StringIO() as file:
                df_place_attributes.to_csv(file, index=False)
                file.seek(0)
                PlaceAttribute.copymanager.from_csv(
                    file,
                    drop_constraints=False, drop_indexes=False,
                )

        # upload the capacities
        df_capacities = pd.DataFrame(capacities,
                                     columns=['place_id', 'service_id', 'capacity',
                                              'from_year', 'to_year'])

        existing_capacities = Capacity.objects.filter(place_id__in=place_ids)
        existing_capacities.delete()

        if len(df_capacities):
            with StringIO() as file:
                df_capacities.to_csv(file, index=False)
                file.seek(0)
                Capacity.copymanager.from_csv(
                    file,
                    drop_constraints=False, drop_indexes=False,
                )
        self.logger.info(f'{len(df_places)} Einträge bearbeitet')
        if n_new > 0:
            self.logger.info(f'davon {n_new} als neue Orte hinzugefügt')
            self.logger.info('ACHTUNG: Für die neuen Orte muss die '
                             'Erreichbarkeit neu berechnet werden!')

    def geocode(self, geocoder: BKGGeocoder, place_row: dict) -> Tuple[float, float]:
        kwargs = {
            'ort': place_row.Ort,
            'plz': place_row.PLZ,
            'strasse': place_row.Straße,
            'haus': place_row.Hausnummer,
        }
        res = geocoder.query(max_retries=2, **kwargs)
        if res.status_code != 200:
            return
        features = res.json()['features']
        if not features:
            return
        return features[0]['geometry']['coordinates']

import os
from collections import OrderedDict
import pandas as pd

from django.conf import settings
from rest_framework import serializers

from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl import styles
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.worksheet.datavalidation import DataValidation

from datentool_backend.user.models import (Infrastructure,
                                           InfrastructureAccess,
                                           PlaceField,
                                           )

from datentool_backend.area.models import MapSymbol, FieldTypes, FClass
from datentool_backend.area.serializers import MapSymbolSerializer
from datentool_backend.infrastructure.serializers import PlaceFieldInfraSerializer


class InfrastructureAccessSerializer(serializers.ModelSerializer):
    class Meta:
        model = InfrastructureAccess
        fields = ('infrastructure', 'profile', 'allow_sensitive_data')
        read_only_fields = ('id', 'infrastructure', )


class InfrastructureSerializer(serializers.ModelSerializer):
    symbol = MapSymbolSerializer(allow_null=True, required=False)
    accessible_by = InfrastructureAccessSerializer(
        many=True, source='infrastructureaccess_set', required=False)
    place_fields = PlaceFieldInfraSerializer(many=True, source='placefield_set',
                                             required=False, read_only=True)

    class Meta:
        model = Infrastructure
        fields = ('id', 'name', 'description',
                  'editable_by', 'accessible_by',
                  'order', 'symbol', 'place_fields')

    def create(self, validated_data):
        symbol_data = validated_data.pop('symbol', None)
        symbol = MapSymbol.objects.create(**symbol_data) \
            if symbol_data else None

        editable_by = validated_data.pop('editable_by', [])
        accessible_by = validated_data.pop('infrastructureaccess_set', [])
        instance = Infrastructure.objects.create(
            symbol=symbol, **validated_data)
        instance.editable_by.set(editable_by)
        instance.accessible_by.set([a['profile'] for a in accessible_by])
        for profile_access in accessible_by:
            infrastructure_access = InfrastructureAccess.objects.get(
                infrastructure=instance,
                profile=profile_access['profile'])
            infrastructure_access.allow_sensitive_data = profile_access['allow_sensitive_data']
            infrastructure_access.save()

        return instance

    def update(self, instance, validated_data):
        # symbol is nullable
        update_symbol = 'symbol' in validated_data
        symbol_data = validated_data.pop('symbol', None)

        editable_by = validated_data.pop('editable_by', None)
        accessible_by = validated_data.pop('infrastructureaccess_set', None)
        instance = super().update(instance, validated_data)
        if editable_by is not None:
            instance.editable_by.set(editable_by)
        if accessible_by is not None:
            instance.accessible_by.set([a['profile'] for a in accessible_by])
            for profile_access in accessible_by:
                infrastructure_access = InfrastructureAccess.objects.get(
                    infrastructure=instance,
                    profile=profile_access['profile'])
                infrastructure_access.allow_sensitive_data = profile_access[
                    'allow_sensitive_data']
                infrastructure_access.save()

        if update_symbol:
            symbol = instance.symbol
            if symbol_data:
                if not symbol:
                    symbol = MapSymbol.objects.create(**symbol_data)
                    instance.symbol = symbol
                else:
                    for k, v in symbol_data.items():
                        setattr(symbol, k, v)
                    symbol.save()
            # symbol passed but is None -> intention to set it to null
            else:
                symbol.delete()
                instance.symbol = None

        instance.save()
        return instance


class InfrastructureTemplateSerializer(serializers.Serializer):
    """Serializer for uploading InfrastructureTemplate"""
    excel_file = serializers.FileField()
    drop_constraints = serializers.BooleanField(default=True)

    def create_template(self, pk: int) -> bytes:
        """Create a template file for infrastructure with the given pk"""

        infrastructure = Infrastructure.objects.get(pk=pk)

        excel_filename = 'Standorte_und_Kapazitäten.xlsx'

        sn_classifications = 'Klassifizierungen'

        columns = {'Name': 'So werden die Einrichtungen auf den Karten beschriftet. '\
                   'Jeder Standort muss einen Namen haben, den kein anderer Standort trägt.',
                  'Straße': 'Postalisch korrekte Schreibweise',
                  'Hausnummer': 'Zahl, ggf. mit Buchstaben (7b) oder 11-15',
                  'PLZ': 'fünfstellig',
                  'Ort': 'Postalisch korrekte Schreibweise',
                  'Lon': 'Längengrad, in WGS84',
                  'Lat': 'Breitengrad, in WGS84',}

        dv_01 = DataValidation(type="whole",
                        operator="between",
                        formula1=0,
                        formula2=1,
                        allow_blank=True)
        dv_01.error = 'Nur 0 oder 1 erlaubt'
        dv_01.title = 'Ungültige 0/1-Werte'

        dv_float = DataValidation(type="decimal",
                                allow_blank=True)
        dv_float.error = 'Nur Zahlen erlaubt'
        dv_float.title = 'Ungültige Zahlen-Werte'

        dv_pos_float = DataValidation(type="decimal",
                                operator="greaterThanOrEqual",
                                formula1=0,
                                allow_blank=True)
        dv_pos_float.error = 'Nur Zahlen >= 0 erlaubt'
        dv_pos_float.title = 'Ungültige positive Zahlen-Werte'

        validations = {}
        classifications = OrderedDict()
        col_no_classifications = 2

        df_places = pd.DataFrame(columns=pd.Index(columns.items()))
        col_no = len(df_places.columns) + 1

        for place_field in infrastructure.placefield_set.all():
            col_no += 1
            col = place_field.name
            if place_field.field_type.ftype == FieldTypes.STRING:
                description = f'Nutzerdefinierte Spalte (Text)'
            elif place_field.field_type.ftype == FieldTypes.NUMBER:
                description = f'Nutzerdefinierte Spalte (Zahl)'
                validations[col_no] = dv_float
            else:
                description = f'Nutzerdefinierte Spalte (Klassifizierte Werte)'
                cf_colno = classifications.get(place_field.field_type.name)

                if cf_colno:
                    df, cn = cf_colno
                else:
                    df = pd.DataFrame(
                        place_field.field_type.fclass_set.values('order', 'value'))\
                        .set_index('order')\
                        .rename(columns={'value': place_field.field_type.name,})

                    cn = col_no_classifications
                    classifications[place_field.field_type.name] = (df, cn)
                    col_no_classifications += 1

                # list validator
                letter = get_column_letter(cn)
                dv_range = f'={quote_sheetname(sn_classifications)}!${letter}$2:${letter}$9999'
                dv = DataValidation(type='list', formula1=dv_range, allow_blank=True)
                dv.error = 'Nur Werte aus Liste erlaubt'
                dv.title = f'Liste für {place_field.field_type.name}'
                validations[col_no] = dv

            df_places = pd.concat([df_places, pd.DataFrame(columns=pd.Index([(col, description)]))],
                           axis=1)

        for service in infrastructure.service_set.all():
            col_no += 1
            if service.has_capacity:
                col = f'Kapazität für Leistung {service.name}'
                description = service.capacity_plural_unit
                dv = dv_pos_float
            else:
                col = f'Bietet Leistung {service.name} an'
                description = '1 wenn ja, sonst 0'
                dv = dv_01

            df_places = pd.concat([df_places, pd.DataFrame(columns=pd.Index([(col, description)]))],
                           axis=1)
            validations[col_no] = dv

        if classifications:df_classification = pd.concat(
            [c[0] for c in classifications.values()],
            axis=1)
        else:
            df_classification = pd.DataFrame()


        fn = os.path.join(settings.MEDIA_ROOT, excel_filename)
        sheetname = 'Standorte und Kapazitäten'
        with pd.ExcelWriter(fn, engine='openpyxl') as writer:
            df_meta = pd.DataFrame({'value': [pk]},
                                   index=pd.Index(['infrastructure'], name='key'))
            df_meta.to_excel(writer, sheet_name='meta')

            df_classification.to_excel(writer, sheet_name=sn_classifications)

            df_places.to_excel(writer, sheet_name=sheetname, freeze_panes=(2, 2))

            ws: Worksheet = writer.sheets.get(sheetname)

            row12 = ws[1:2]
            for row in row12:
                for cell in row:
                    cell.alignment = styles.Alignment(horizontal='center', wrap_text=True)

            dv = DataValidation(type="decimal",
                                operator="between",
                                formula1=0,
                                formula2=90,
                                allow_blank=True)

            dv.error ='Koordinaten müssen in WGS84 angegeben werden und zwischen 0 und 90 liegen'
            dv.errorTitle = 'Ungültige Koordinaten'
            ws.add_data_validation(dv)
            ws.add_data_validation(dv_01)
            ws.add_data_validation(dv_float)
            ws.add_data_validation(dv_pos_float)
            dv.add('G3:H999999')
            for col_no, dv in validations.items():
                letter = get_column_letter(col_no)
                cell_range = f'{letter}3:{letter}999999'
                dv.add(cell_range)
                if not dv in ws.data_validations:
                    ws.add_data_validation(dv)

            ws.column_dimensions['A'] = ColumnDimension(ws, index='A', hidden=True)
            ws.column_dimensions['B'] = ColumnDimension(ws, index='B', width=30)
            for col_no in range(3, len(df_places.columns) + 2):
                col = get_column_letter(col_no)
                ws.column_dimensions[col] = ColumnDimension(ws, index=col, width=16)

        content = open(fn, 'rb').read()
        return content

    def read_excel_file(self, request) -> pd.DataFrame:
        """read excelfile and return a dataframe"""
        excel_file = request.FILES['excel_file']

        df_meta = pd.read_excel(excel_file.file,
                           sheet_name='meta',
                           skiprows=[1])

        df = pd.read_excel(excel_file.file,
                           sheet_name='Standorte und Kapazitäten',
                           skiprows=[1])


        df.rename(columns={'from_stop': 'from_stop_id',
                           'to_stop': 'to_stop_id',}, inplace=True)

        variant = request.data.get('variant')
        df['variant_id'] = int(variant)
        return df

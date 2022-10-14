import os
from io import BytesIO
import pandas as pd
from openpyxl.reader.excel import load_workbook

from matrixconverters.read_ptv import ReadPTVMatrix

from django.urls import reverse
from test_plus import APITestCase

from datentool_backend.api_test import LoginTestCase
from datentool_backend.indicators.models import Stop, MatrixStopStop
from datentool_backend.modes.factories import ModeVariantFactory
from datentool_backend.modes.models import Mode

class StopTemplateTest(LoginTestCase, APITestCase):

    testdata_folder = 'testdata'
    filename_stops = 'Haltestellen.xlsx'
    filename_stops_errors = 'Haltestellen_errors.xlsx'
    filename_rz = 'Reisezeitmatrix_Haltestelle_zu_Haltestelle.xlsx'
    filename_rz_errors = 'Reisezeitmatrix_errors.xlsx'

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.profile.can_edit_basedata = True
        cls.profile.save()
        cls.mode_variant = ModeVariantFactory(label='ÖV-Basis',
                                              mode=Mode.TRANSIT)

    def get_file_path_stops(self, filename_stops: str=None) -> str:
        file_path_stops = os.path.join(os.path.dirname(__file__),
                                    self.testdata_folder,
                                    filename_stops or self.filename_stops)
        return file_path_stops

    def test_request_stop_template(self):
        url = reverse('stops-create-template')
        res = self.post(url, extra={'format': 'json'})
        self.assert_http_200_ok(res)
        wb = load_workbook(BytesIO(res.content))
        self.assertListEqual(wb.sheetnames, ['Haltestellen'])

    def test_upload_stop_template(self):
        """
        test bulk upload stops
        """

        file_path_stops = self.get_file_path_stops()
        file_content = open(file_path_stops, 'rb')
        data = {
            'excel_file': file_content,
            'variant': self.mode_variant.id,
            'sync': True,
        }

        url = reverse('stops-upload-template')
        res = self.client.post(url, data, extra=dict(format='multipart/form-data'))
        self.assert_http_202_accepted(res, msg=res.content)
        # 4 stops should have been uploaded with the correct hst-nr and names
        df = pd.read_excel(file_path_stops, skiprows=[1])

        actual = pd.DataFrame(Stop.objects.values('hstnr', 'name')).set_index('hstnr')
        expected = df[['HstNr', 'HstName']]\
            .rename(columns={'HstNr': 'hstnr', 'HstName': 'name', })\
            .set_index('hstnr')
        pd.testing.assert_frame_equal(actual, expected)

        # assert that without edit_basedata-permissions no upload is possible
        self.profile.can_edit_basedata = False
        self.profile.save()
        res = self.client.post(url, data, extra=dict(format='multipart/form-data'))
        self.assert_http_403_forbidden(res)

    def test_upload_broken_stop_file(self):
        """
        test errors in upload files
        """
        file_path = os.path.join(os.path.dirname(__file__),
                                 self.testdata_folder,
                                 self.filename_stops_errors)
        url = reverse('stops-upload-template')
        data = {
            'excel_file': open(file_path, 'rb'),
            'drop_constraints': False,
            'variant': self.mode_variant.id,
            'sync': True
        }
        res = self.client.post(url, data, extra=dict(format='multipart/form-data'))
        self.assertContains(res,
                            'Haltestellennummer ist nicht eindeutig',
                            status_code=406)
        print(res.content)

    def create_mode_variant(self) -> int:
        return ModeVariantFactory().pk

    def upload_stops(self, filename_stops: str='Haltestellen.xlsx'):
        url = reverse('stops-upload-template')
        data = {
            'excel_file': open(self.get_file_path_stops(filename_stops), 'rb'),
            'drop_constraints': False,
            'variant': self.mode_variant.id,
            'sync':  True
        }
        res = self.client.post(url, data, extra=dict(format='multipart/form-data'))
        self.assert_http_202_accepted(res, msg=res.content)

    def test_request_template_matrixstopstop(self):
        data = {
            'variant': self.mode_variant.id,
        }
        url = reverse('matrixstopstops-create-template')
        res = self.post(url, data=data, extra={'format': 'json'})
        self.assert_http_200_ok(res)
        wb = load_workbook(BytesIO(res.content))
        self.assertListEqual(wb.sheetnames, ['Reisezeit'])

    def test_upload_matrixstopstop_template(self):
        """
        test bulk upload matrix
        """
        for fn_stops, fn_matrix in [
            ('Haltestellen.xlsx', 'Reisezeitmatrix_Haltestelle_zu_Haltestelle.xlsx'),
            ('Haltestellen_HL.xlsx', 'Reisezeimatrix_HL_klein.mtx')
        ]:

            self.upload_stops(fn_stops)
            mode_variant_id = self.mode_variant.id
            file_path = os.path.join(os.path.dirname(__file__),
                                        self.testdata_folder,
                                        fn_matrix)
            file_content = open(file_path, 'rb')
            data = {
                'excel_or_visum_file': file_content,
                'variant': mode_variant_id,
                'drop_constraints': False,
                'sync': True,
            }

            url = reverse('matrixstopstops-upload-template')
            res = self.client.post(url, data, extra=dict(format='multipart/form-data'))
            self.assert_http_202_accepted(res, msg=res.content)

            if fn_matrix.endswith('.xlsx'):
                df = pd.read_excel(file_path, skiprows=[1])
            elif fn_matrix.endswith('mtx'):
                da = ReadPTVMatrix(file_path)
                df = da['matrix'].to_dataframe()
                df = df.loc[df['matrix']<999999]
                df.index.rename(['from_stop', 'to_stop'], inplace=True)
                df.rename(columns={'matrix': 'minutes',}, inplace=True)
                df.reset_index(inplace=True)

            cols = ['id', 'name', 'hstnr']
            df_stops = pd.DataFrame(Stop.objects.filter(variant=mode_variant_id).values(*cols),
                                    columns=cols)\
                .set_index('hstnr')

            df = df\
                .merge(df_stops['id'].rename('from_stop_id'),
                          left_on='from_stop', right_index=True)\
                .merge(df_stops['id'].rename('to_stop_id'),
                          left_on='to_stop', right_index=True)

            actual = pd.DataFrame(
                MatrixStopStop.objects.values('from_stop_id', 'to_stop_id', 'minutes'))\
                .set_index(['from_stop_id', 'to_stop_id'])
            expected = df[['from_stop_id', 'to_stop_id', 'minutes']]\
                .set_index(['from_stop_id', 'to_stop_id'])
            pd.testing.assert_frame_equal(actual, expected, check_dtype=False)

    def test_upload_broken_matrix_file(self):
        """
        test errors in upload files
        """
        self.upload_stops()
        mode_variant_id = self.mode_variant.id
        file_path = os.path.join(os.path.dirname(__file__),
                                 self.testdata_folder,
                                 self.filename_rz_errors)
        url = reverse('matrixstopstops-upload-template')
        data = {
            'excel_or_visum_file' : open(file_path, 'rb'),
            'variant': mode_variant_id,
            'drop_constraints': False,
            'sync': True,
        }
        res = self.client.post(url, data, extra=dict(format='multipart/form-data'), varant=33)
        self.assertContains(res,
                            'Haltestelle nicht in Haltestellennummern',
                            status_code=406)
        print(res.content)

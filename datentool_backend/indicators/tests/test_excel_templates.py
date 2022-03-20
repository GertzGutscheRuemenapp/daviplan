import os
from io import BytesIO
from openpyxl.reader.excel import load_workbook
import pandas as pd

from django.urls import reverse
from test_plus import APITestCase
from datentool_backend.api_test import LoginTestCase
from datentool_backend.indicators.models import Stop, MatrixStopStop
from datentool_backend.modes.factories import ModeVariantFactory


class StopTemplateTest(LoginTestCase, APITestCase):

    testdata_folder = 'testdata'
    filename_stops = 'Haltestellen.xlsx'
    filename_stops_errors = 'Haltestellen_errors.xlsx'
    filename_rz = 'Reisezeitmatrix_Haltestelle_zu_Haltestelle.xlsx'
    filename_rz_errors = 'Reisezeitmatrix_errors.xlsx'

    @property
    def file_path_stops(self) -> str:
        file_path_stops = os.path.join(os.path.dirname(__file__),
                                    self.testdata_folder,
                                    self.filename_stops)
        return file_path_stops

    def test_request_stop_template(self):
        url = reverse('stops-download-template')
        res = self.get_check_200(url)
        wb = load_workbook(BytesIO(res.content))
        self.assertListEqual(wb.sheetnames, ['Haltestellen'])

    def test_upload_stop_template(self):
        """
        test bulk upload stops
        """
        file_content = open(self.file_path_stops, 'rb')
        data = {
            'excel_file' : file_content,
        }

        url = reverse('stops-upload-template')
        res = self.client.post(url, data, extra=dict(format='multipart/form-data'))
        self.assert_http_403_forbidden(res)
        file_content.seek(0)
        self.profile.can_edit_basedata = True
        self.profile.save()
        res = self.client.post(url, data, extra=dict(format='multipart/form-data'))
        self.assert_http_202_accepted(res, msg=res.content)
        # 4 stops should have been uploaded with the correct hst-nr and names
        df = pd.read_excel(self.file_path_stops, skiprows=[1])

        actual = pd.DataFrame(Stop.objects.values('id', 'name')).set_index('id')
        expected = df[['Nr', 'Name']].rename(columns={'Nr': 'id','Name': 'name',}).set_index('id')
        pd.testing.assert_frame_equal(actual, expected)

    def test_upload_broken_stop_file(self):
        """
        test errors in upload files
        """
        file_path = os.path.join(os.path.dirname(__file__),
                                 self.testdata_folder,
                                 self.filename_stops_errors)
        url = reverse('stops-upload-template')
        data = {
            'excel_file' : open(file_path, 'rb'),
            'drop_constraints': False,
        }
        self.profile.can_edit_basedata = True
        self.profile.save()
        res = self.client.post(url, data, extra=dict(format='multipart/form-data'))
        self.assertContains(res,
                            'Haltestellennummer ist nicht eindeutig',
                            status_code=406)
        print(res.content)

    def create_mode_variant(self) -> int:
        return ModeVariantFactory().pk

    def upload_stops(self):
        self.profile.can_edit_basedata = True
        self.profile.save()
        url = reverse('stops-upload-template')
        data = {
            'excel_file' : open(self.file_path_stops, 'rb'),
            'drop_constraints': False,
        }
        res = self.client.post(url, data, extra=dict(format='multipart/form-data'))
        self.assert_http_202_accepted(res, msg=res.content)

    def test_request_template_matrixstopstop(self):
        url = reverse('matrixstopstops-download-template')
        res = self.get_check_200(url)
        wb = load_workbook(BytesIO(res.content))
        self.assertListEqual(wb.sheetnames, ['Reisezeit'])

    def test_upload_matrixstopstop_template(self):
        """
        test bulk upload matrix
        """
        self.upload_stops()
        mode_variant_id = self.create_mode_variant()
        file_path = os.path.join(os.path.dirname(__file__),
                                    self.testdata_folder,
                                    self.filename_rz)
        file_content = open(file_path, 'rb')
        data = {
            'excel_file' : file_content,
            'variant': mode_variant_id,
            'drop_constraints': False,
        }

        url = reverse('matrixstopstops-upload-template')
        self.profile.can_edit_basedata = True
        self.profile.save()
        res = self.client.post(url, data, extra=dict(format='multipart/form-data'))
        self.assert_http_202_accepted(res, msg=res.content)

        df = pd.read_excel(file_path, skiprows=[1])

        actual = pd.DataFrame(
            MatrixStopStop.objects.values('from_stop', 'to_stop', 'minutes'))\
            .set_index(['from_stop', 'to_stop'])
        expected = df[['from_stop', 'to_stop', 'minutes']].set_index(['from_stop', 'to_stop'])
        pd.testing.assert_frame_equal(actual, expected, check_dtype=False)

    def test_upload_broken_matrix_file(self):
        """
        test errors in upload files
        """
        self.upload_stops()
        mode_variant_id = self.create_mode_variant()
        file_path = os.path.join(os.path.dirname(__file__),
                                 self.testdata_folder,
                                 self.filename_rz_errors)
        url = reverse('matrixstopstops-upload-template')
        data = {
            'excel_file' : open(file_path, 'rb'),
            'variant': mode_variant_id,
            'drop_constraints': False,
        }
        self.profile.can_edit_basedata = True
        self.profile.save()
        res = self.client.post(url, data, extra=dict(format='multipart/form-data'), varant=33)
        self.assertContains(res,
                            'Haltestelle nicht in Haltestellennummern',
                            status_code=406)
        print(res.content)

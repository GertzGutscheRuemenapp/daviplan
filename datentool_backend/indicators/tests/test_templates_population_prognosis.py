import os
from io import BytesIO
import pandas as pd
from openpyxl.reader.excel import load_workbook

from django.urls import reverse
from test_plus import APITestCase

from datentool_backend.indicators.tests.setup_testdata import CreateTestdataMixin
from datentool_backend.api_test import LoginTestCase

from datentool_backend.site.factories import YearFactory
from datentool_backend.demand.factories import GenderFactory, AgeGroup
from datentool_backend.area.factories import (AreaLevelFactory,
                                              AreaFactory,
                                              AreaFieldFactory,
                                              FieldTypeFactory,
                                              FieldTypes)


from datentool_backend.population.factories import (PopulationRasterFactory,
                                                    PopulationFactory,
                                                    PrognosisFactory,
                                                    PopulationEntry)


class PopulationTemplateTest(LoginTestCase, APITestCase, CreateTestdataMixin):

    testdata_folder = 'testdata'

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.profile.can_edit_basedata = True
        cls.profile.save()

        cls.genders = [GenderFactory(name=n) for n in ['männlich', 'weiblich']]
        AgeGroup.objects.create(from_age=0, to_age=17)
        AgeGroup.objects.create(from_age=18, to_age=64)
        AgeGroup.objects.create(from_age=65, to_age=120)

        cls.age_groups = AgeGroup.objects.all()

        cls.popraster = PopulationRasterFactory(default=True)

        cls.years = [YearFactory(year=y, is_default=(y==2022))
                     for y in range(2015, 2023)]
        for year in cls.years:
            PopulationFactory(popraster=cls.popraster, year=year, prognosis=None)

        cls.prognosis = PrognosisFactory(id=1, name='Trendentwicklung', is_default=True)
        cls.prognosis_years = [YearFactory(year=y) for y in range(2025, 2045, 5)]
        for year in cls.prognosis_years:
            PopulationFactory(popraster=cls.popraster, year=year, prognosis=cls.prognosis)

        cls.area_level = AreaLevelFactory(id=1, name='Gemeinden',
                                          is_default_pop_level=True)

        str_field = FieldTypeFactory(ftype=FieldTypes.STRING)

        cls.gen = AreaFieldFactory(field_type=str_field,
                                   area_level=cls.area_level,
                                   name='gen',
                                   is_label=True)
        cls.rs = AreaFieldFactory(field_type=str_field,
                                  area_level=cls.area_level,
                                  name='rs',
                                  is_key=True)

        areas = [['010030000000', 'Lübeck'],
                 ['010530090090', 'Mölln'],
                 ['010530100100', 'Ratzeburg'],
                 ['010535308008', 'Behlendorf'],
                 ['010535308009', 'Berkenthin'],
                 ]

        for rs, gen in areas:
            area = AreaFactory(area_level=cls.area_level)
            area.attributes = {'rs': rs, 'gen': gen,}

    def test_create_population_template(self):
        """Create templates for Population"""
        url = reverse('populationentries-create-template')
        years = [y.year for y in self.years]
        res = self.post(url, data={'area_level': self.area_level.pk,
                                   'years': years,
                                   },
                             extra=dict(format='json')
                             )
        self.assert_http_200_ok(res)
        wb = load_workbook(BytesIO(res.content))
        self.assertSetEqual(set(wb.sheetnames),
                            set([f'{y}' for y in years]))

    def test_create_prognosis_template(self):
        """Create templates for Prognosis"""
        url = reverse('populationentries-create-template')
        years = [y.year for y in self.prognosis_years]
        res = self.post(url, data={'area_level': self.area_level.pk,
                                   'years': years,
                                   'prognosis': self.prognosis.pk,},
                             extra=dict(format='json')
                             )
        self.assert_http_200_ok(res)
        wb = load_workbook(BytesIO(res.content))
        self.assertSetEqual(set(wb.sheetnames),
                            set([f'{y}' for y in years]))

    def test_upload_prognosis_template(self):
        """
        test bulk upload population_entries
        """

        # upload excel-file
        file_names_popdata = ['Einwohnerrealdaten.xlsx', 'Prognosedaten.xlsx']
        for file_name_popdata in file_names_popdata:
            file_path_popdata = os.path.join(os.path.dirname(__file__),
                                            self.testdata_folder,
                                            file_name_popdata)
            file_content = open(file_path_popdata, 'rb')
            data = {
                'excel_file' : file_content,
                'prognosis': 1,
            }

            url = reverse('populationentries-upload-template')
            res = self.client.post(url, data,
                                   extra=dict(format='multipart/form-data'))
            self.assert_http_202_accepted(res, msg=res.content)

            df_actual = pd.DataFrame(PopulationEntry.objects\
                         .values('population__year__year','value'))\
                .groupby('population__year__year').sum()

            wb = load_workbook(file_path_popdata)
            meta = wb['meta']
            n_years = meta['B2'].value
            years = [meta.cell(3, n).value for n in range(2, n_years + 2)]
            for y in years:
                total_pop = pd.read_excel(file_path_popdata,
                                          sheet_name=str(y),
                                          header=[1, 3],
                                          skiprows=[4],
                                          index_col=[0, 1])\
                    .sum().sum()
                self.assertAlmostEqual(total_pop, df_actual.loc[y, 'value'])

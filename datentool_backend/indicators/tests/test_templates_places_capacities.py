import os
from io import BytesIO
import warnings
from unittest import skipIf

import pandas as pd
from openpyxl.reader.excel import load_workbook

from django.urls import reverse
from test_plus import APITestCase

from datentool_backend.api_test import LoginTestCase
from datentool_backend.utils.test_utils import no_connection
from datentool_backend.utils.bkg_geocoder import BASE_URL
from datentool_backend.site.models import SiteSetting
from datentool_backend.utils.crypto import encrypt

from datentool_backend.infrastructure.factories import (InfrastructureFactory,
                                                        ServiceFactory,
                                                        )
from datentool_backend.places.factories import (Place,
                                                CapacityFactory,
                                                PlaceFieldFactory,
                                                PlaceFactory,
                                                )
from datentool_backend.area.models import FieldTypes
from datentool_backend.area.factories import FClassFactory, FieldTypeFactory


class InfrastructureTemplateTest(LoginTestCase, APITestCase):

    testdata_folder = 'testdata'

    @classmethod
    def setUpTestData(cls):
        super().setUpTestData()
        cls.profile.can_edit_basedata = True
        cls.profile.save()

        site_settings = SiteSetting.load()
        bkg_key = os.environ.get('BKG_KEY', 'aaa-bbb-ccc')
        site_settings.bkg_password = encrypt(bkg_key)
        site_settings.save()

        cls.infra = InfrastructureFactory(id=1)
        cls.service1 = ServiceFactory(name='Schulen',
                                      has_capacity=False,
                                      infrastructure=cls.infra)
        cls.service2 = ServiceFactory(name='Schulplätze',
                                      has_capacity=True,
                                      infrastructure=cls.infra)

        #cls.str_field = PlaceFieldFactory(infrastructure=cls.infra,
                                          #field_type__name='PlaceName',
                                          #field_type__ftype=FieldTypes.STRING,
                                          #name='Bezeichnung',
                                          #is_label=True)
        cls.hnr_field = PlaceFieldFactory(infrastructure=cls.infra,
                                          field_type__name='Hnr',
                                          name='Hausnummer',
                                          field_type__ftype=FieldTypes.STRING)
        cls.num_field = PlaceFieldFactory(infrastructure=cls.infra,
                                          field_type__name='FloatType',
                                          name='Gewichtung',
                                          field_type__ftype=FieldTypes.NUMBER,
                                          unit='Punkte')

        cl_ft1 = FieldTypeFactory(ftype=FieldTypes.CLASSIFICATION, name='Ziffern')
        cl_ft2 = FieldTypeFactory(ftype=FieldTypes.CLASSIFICATION, name='Buchstaben')
        cls.cla_field = PlaceFieldFactory(infrastructure=cls.infra,
                                          name='EinsZweiDrei',
                                          field_type=cl_ft1)

        cv11 = FClassFactory(ftype=cls.cla_field.field_type, value='Eins', order=1)
        cv12 = FClassFactory(ftype=cls.cla_field.field_type, value='Zwei', order=2)
        cv13 = FClassFactory(ftype=cls.cla_field.field_type, value='Drei', order=3)

        cls.cla_field2 = PlaceFieldFactory(infrastructure=cls.infra,
                                           name='AABB',
                                           field_type=cl_ft2)

        cv21 = FClassFactory(ftype=cls.cla_field2.field_type, value='AA', order=1)
        cv22 = FClassFactory(ftype=cls.cla_field2.field_type, value='BB', order=2)

        place1 = PlaceFactory(infrastructure=cls.infra,
                              name='Place1',
                              attributes={
            cls.num_field.name: 1234,
            cls.hnr_field.name: '44b',
            cls.cla_field.name: 'Zwei',
            cls.cla_field2.name: 'BB',
        })
        place2 = PlaceFactory(infrastructure=cls.infra,
                              name='Place2',
                              attributes={
            cls.num_field.name: 567,
            cls.hnr_field.name: '33',
            cls.cla_field.name: 'Drei',
        })

        CapacityFactory(service=cls.service1, place=place1, capacity=1)
        CapacityFactory(service=cls.service1, place=place2, capacity=0)
        CapacityFactory(service=cls.service2, place=place1, capacity=55.5)
        CapacityFactory(service=cls.service2, place=place2, capacity=0)

    def test_create_infrastructure_template(self):
        url = reverse('places-create-template')
        res = self.post(url, data={'infrastructure': self.infra.pk,},
                        extra={'format': 'json'})
        self.assert_http_200_ok(res)
        wb = load_workbook(BytesIO(res.content))
        self.assertSetEqual(set(wb.sheetnames),
                            {'Standorte und Kapazitäten', 'meta', 'Klassifizierungen'})
        df = pd.read_excel(BytesIO(res.content), sheet_name='Klassifizierungen')
        expected = pd.DataFrame({'order': [1, 2, 3],
                                 'Ziffern': ['Eins', 'Zwei', 'Drei'],
                                 'Buchstaben': ['AA', 'BB', pd.NA], })
        pd.testing.assert_frame_equal(df, expected)
        # check if the unit was used for column "Gewichtung"
        # get the values and unpivot the data
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", category=UserWarning)
            df = pd.read_excel(BytesIO(res.content),
                               sheet_name='Standorte und Kapazitäten')
        self.assertEqual(df.loc[0, 'Gewichtung'],
                         'Nutzerdefinierte Spalte (Punkte)')

    def test_upload_broken_place_template(self):
        """
        test bulk upload places and capacities
        """
        # delete Place1
        Place.objects.get(name='Place1').delete()

        # upload excel-file
        file_name_places = 'Standorte_und_Kapazitäten_mod.xlsx'
        file_path_places = os.path.join(os.path.dirname(__file__),
                                        self.testdata_folder,
                                        file_name_places)
        file_content = open(file_path_places, 'rb')
        data = {
            'excel_file' : file_content,
            'infrastructure': self.infra.pk,
        }

        url = reverse('places-upload-template')
        res = self.client.post(url, data,
                               extra=dict(format='multipart/form-data'))
        # modified excel-Testfile with non-matching-classifications should
        # raise a 406-response
        self.assert_http_406_not_acceptable(res, msg=res.content)

    def test_upload_place_template(self):

        # delete Place1
        Place.objects.get(name='Place1').delete()
        place2 = Place.objects.get(name='Place2')


        # create a third place, that should be deleted
        place3 = PlaceFactory(infrastructure=place2.infrastructure)

        # upload excel-file
        file_name_places = 'Standorte_und_Kapazitäten.xlsx'
        file_path_places = os.path.join(os.path.dirname(__file__),
                                        self.testdata_folder,
                                        file_name_places)
        file_content = open(file_path_places, 'rb')
        data = {
            'excel_file': file_content,
            'infrastructure': self.infra.pk,
        }

        url = reverse('places-upload-template')
        res = self.client.post(url, data,
                               extra=dict(format='multipart/form-data'))


        self.assert_http_202_accepted(res, msg=res.content)

        # get the values and unpivot the data
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", category=UserWarning)
            df = pd.read_excel(file_path_places, sheet_name='Standorte und Kapazitäten',
                               skiprows=[1]).set_index('place_id')

        places = Place.objects.filter(infrastructure=self.infra)
        place_names = places.values_list('name', flat=True)
        assert set(df['Name']).issubset(place_names),\
               'place names in excel_file are not uploaded correcty'

        for place_id, place_row in df.iterrows():
            place = Place.objects.get(infrastructure=self.infra, name=place_row['Name'])
            coords = place.geom.transform(4326, clone=True).coords
            for i, c in enumerate(['Lon', 'Lat']):
                self.assertAlmostEqual(coords[i], place_row.loc[c])
            for place_attr in place.placeattribute_set.all():
                value = place_row.get(place_attr.field.name)
                if place_attr.field.field_type.ftype in [FieldTypes.STRING,
                                                         FieldTypes.CLASSIFICATION]:
                    value = str(value)
                self.assertAlmostEqual(place_attr.value, value)
            for capacity in place.capacity_set.all():
                service = capacity.service
                if service.has_capacity:
                    col = f'Kapazität für Leistung {service.name}'
                else:
                    col = f'Bietet Leistung {service.name} an'
                value = place_row.loc[col]
                self.assertAlmostEqual(capacity.capacity, value)


    @skipIf(no_connection(BASE_URL), 'BKG Geocoding services not available')
    def test_geocode_places(self):
        # delete all existing places
        Place.objects.all().delete()

        # upload excel-file
        file_name_places = 'Standorte_und_Kapazitäten_ohne_Koordinaten.xlsx'
        file_path_places = os.path.join(os.path.dirname(__file__),
                                        self.testdata_folder,
                                        file_name_places)
        file_content = open(file_path_places, 'rb')
        data = {
            'excel_file': file_content,
            'infrastructure': self.infra.pk,
        }

        url = reverse('places-upload-template')
        res = self.client.post(url, data,
                               extra=dict(format='multipart/form-data'))


        self.assert_http_202_accepted(res, msg=res.content)

        # get the values and unpivot the data
        with warnings.catch_warnings():
            warnings.simplefilter("ignore", category=UserWarning)
            df = pd.read_excel(file_path_places, sheet_name='Standorte und Kapazitäten',
                               skiprows=[1]).set_index('place_id')

        places = Place.objects.filter(infrastructure=self.infra).order_by('id')
        place_names = places.values_list('name', flat=True)
        # only Place1 and Place3 can be geocoded
        self.assertQuerysetEqual(place_names, ['Place1', 'Place3'])
        # the other rows should have been deleted/skipped, because the geocoding failed

        df_places_with_valid_adresses = df.iloc[[0, 2]]

        # check the geocoded adresses
        for place_id, place_row in df_places_with_valid_adresses.iterrows():
            place = Place.objects.get(infrastructure=self.infra, name=place_row['Name'])
            lon, lat = place.geom.transform(4326, clone=True).coords
            self.assertAlmostEqual(lon, 10.56744104, places=6)
            self.assertAlmostEqual(lat, 54.01146757, places=6)

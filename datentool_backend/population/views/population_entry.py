from typing import Dict
import warnings
import os
from tempfile import mktemp
import pandas as pd
import datetime

from openpyxl.reader.excel import load_workbook

from drf_spectacular.utils import (extend_schema,
                                   inline_serializer)
from djangorestframework_camel_case.parser import CamelCaseMultiPartParser
from rest_framework import viewsets
from rest_framework.decorators import action
from django.core.exceptions import BadRequest

from datentool_backend.utils.excel_template import (ColumnError,
                                                    write_template_df,
                                                    ExcelTemplateMixin)

from datentool_backend.utils.permissions import (
    HasAdminAccessOrReadOnly, HasAdminAccess, CanEditBasedata, IsDemoUser)
from datentool_backend.utils.pop_aggregation import (
    aggregate_many,
    disaggregate_population)
from datentool_backend.utils.raw_delete import delete_chunks

from datentool_backend.population.models import (
    Population,
    Gender,
    PopulationEntry,
    Year
)

from datentool_backend.demand.models import AgeGroup
from datentool_backend.utils.serializers import drop_constraints
from datentool_backend.population.serializers import (
    PopulationEntrySerializer,
    PopulationTemplateSerializer,
    prognosis_id_serializer,
    area_level_id_serializer,
    years_serializer,
    get_area_level_key,
)
from datentool_backend.area.models import Area, AreaLevel

import logging

logger = logging.getLogger('population')


class PopulationEntryViewSet(ExcelTemplateMixin, viewsets.ModelViewSet):
    queryset = PopulationEntry.objects.all()
    serializer_class = PopulationEntrySerializer
    serializer_action_classes = {'create_template': PopulationTemplateSerializer,
                                 'upload_template': PopulationTemplateSerializer,
                                 }
    permission_classes = [HasAdminAccessOrReadOnly | CanEditBasedata]
    filterset_fields = ['population']

    @extend_schema(description='Upload Population Template',
                   request=inline_serializer(
                       name='PopulationUploadTemplateSerializer',
                       fields={
                           'area_level': area_level_id_serializer,
                           'years': years_serializer,
                           'prognosis': prognosis_id_serializer,
                           'drop_constraints': drop_constraints
                       }
                       ),
                   )
    @action(methods=['POST'], detail=False,
            permission_classes=[IsDemoUser | HasAdminAccess | CanEditBasedata])
    def create_template(self, request, **kwargs):
        area_level_id = request.data.get('area_level')
        prognosis_id = request.data.get('prognosis')
        years = request.data.get('years')
        if not years:
            msg = f'Kein Jahr ausgewählt, daher kann kein Template erzeugt werden.'
            logger.error(msg)
            raise BadRequest(msg)
        return super().create_template(request,
                                       area_level_id=area_level_id,
                                       years=years,
                                       prognosis_id=prognosis_id,
                                       )

    @action(methods=['POST'], detail=False,
            permission_classes=[HasAdminAccess | CanEditBasedata],
            parser_classes=[CamelCaseMultiPartParser])
    def upload_template(self, request):
        """Upload the filled out Stops-Template"""
        return super().upload_template(request)

    def get_read_excel_params(self, request) -> Dict:
        params = dict()
        logger.info('Lese Eingangsdatei')
        io_file = request.FILES['excel_file']
        ext = os.path.splitext(io_file.name)[-1]
        fp = mktemp(suffix=ext)
        with open(fp, 'wb') as f:
            f.write(io_file.file.read())
        params['excel_filepath'] = fp
        params['prognosis_id'] = request.data.get('prognosis')
        return params

    @staticmethod
    def process_excelfile(logger,
                          excel_filepath,
                          prognosis_id,
                          drop_constraints=False,
                          ):
        # read excelfile
        logger.info('Lese Excel-Datei')
        df = read_excel_file(excel_filepath, prognosis_id)
        os.remove(excel_filepath)

        queryset = PopulationEntry.objects.filter(population__prognosis=prognosis_id)
        delete_chunks(queryset, logger)

        # write_df
        write_template_df(df, PopulationEntry, logger, drop_constraints=drop_constraints)
        # postprocess (optional)
        post_processing(dataframe=df, drop_constraints=drop_constraints, logger=logger)


def read_excel_file(excel_filepath, prognosis_id) -> pd.DataFrame:
    """read excelfile and return a dataframe"""

    columns = ['population_id', 'area_id', 'gender_id',
               'age_group_id', 'value']
    df = pd.DataFrame(columns=columns)

    wb = load_workbook(excel_filepath)
    area_level = AreaLevel.objects.get(is_default_pop_level=True)

    areas = Area.annotated_qs(area_level)
    key_attr = get_area_level_key(logger, area_level)
    df_areas = pd.DataFrame(areas.values(key_attr, 'id'))\
        .set_index(key_attr)\
        .rename(columns={'id': 'area_id', })

    df_genders = pd.DataFrame(Gender.objects.values('id', 'name'))\
        .set_index('name')\
        .rename(columns={'id': 'gender_id', })
    df_agegroups = pd.DataFrame([[ag.id, ag.name] for
                                 ag in AgeGroup.objects.all()],
                                columns=['age_group_id', 'Altersgruppe'])\
        .set_index('Altersgruppe')

    years = []
    for sn in wb.sheetnames:
        # ignore the former existing hidden meta sheet that was removed
        if sn == 'meta':
            continue
        sheet = wb[sn]
        year = sheet.cell(1, 2).value
        error_msg = ''
        try:
            if int(sn) != year:
                error_msg = (f'Der Name des Blatts "{sn}" stimmt nicht mit dem'
                             f' dort angegebenen Jahr "{year}" überein')
        except ValueError:
            error_msg = (f'Der Name des Blatts "{sn}" entspricht '
                         'keinem gültigen Jahr')
        if error_msg:
            logger.error(error_msg)
            raise Exception(error_msg)
        years.append(year)
    cur_year = datetime.date.today().year
    if prognosis_id is None:
        future_years = [y for y in years if y > cur_year]
        if future_years:
            msg = ('Die Datei enthält Jahre, die in der Zukunft liegen: '
                   f'{future_years}. Laden Sie Daten für zukünftige Jahre bitte'
                   ' als Prognosen unter "Zukunft" hoch.')
            logger.error(msg)
            raise Exception(msg)
    populations = []
    for y in years:
        year, created = Year.objects.get_or_create(year=y)
        population, created = Population.objects.get_or_create(
            prognosis_id=prognosis_id,
            year=year,
        )
        try:
            # get the values and unpivot the data
            with warnings.catch_warnings():
                warnings.simplefilter("ignore", category=UserWarning)
                df_pop = pd.read_excel(excel_filepath,
                                       sheet_name=str(y),
                                       header=[1, 2, 3, 4],
                                       dtype={key_attr: object, },
                                       index_col=[0, 1])
            df_pop = df_pop\
                .stack(level=[0, 1, 2, 3])\
                .reset_index()
            df_pop = df_pop.drop(['gender_id', 'age_group_id'], axis=1)
            df_pop = df_pop\
                .merge(df_areas, left_on=key_attr, right_index=True)\
                .merge(df_genders, left_on='Geschlecht', right_index=True)\
                .merge(df_agegroups, left_on='Altersgruppe', right_index=True)

            df_pop.rename(columns={0: 'value',}, inplace=True)
            df_pop['population_id'] = population.id
        except KeyError as e:
            msg = f'Spalte {e} wurde nicht gefunden.'
            logger.error(msg)
            raise ColumnError(msg)
        populations.append(population)
        df = pd.concat([df, df_pop[columns]])

    return df


def post_processing(dataframe, drop_constraints=False,
                    logger=logging.getLogger('population')):
    populations = Population.objects.filter(
        id__in=dataframe['population_id'].unique())
    logger.info('Disaggregiere Bevölkerungsdaten')
    for i, population in enumerate(populations):
        disaggregate_population(population, use_intersected_data=True,
                                drop_constraints=drop_constraints)
        logger.info(f'{i + 1:n}/{len(populations):n} Jahren bearbeitet')
    logger.info('Aggregiere Bevölkerungsdaten')
    aggregate_many(AreaLevel.objects.all(), populations,
                   drop_constraints=drop_constraints)

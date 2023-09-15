import os
from typing import List
import pandas as pd

from django.core.exceptions import BadRequest
from django.conf import settings
from rest_framework import serializers

from openpyxl.utils import get_column_letter
from openpyxl import styles
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.dimensions import ColumnDimension, RowDimension
from openpyxl.worksheet.datavalidation import DataValidation

from datentool_backend.utils.processes import ProcessScope
from datentool_backend.area.models import AreaLevel, Area, AreaField
from datentool_backend.demand.models import Gender, AgeGroup
from datentool_backend.population.models import (Population,
                                                 Prognosis,
                                                 PopulationEntry)

import logging


area_level_id_serializer = serializers.PrimaryKeyRelatedField(
    queryset=AreaLevel.objects.all(),
    help_text='area_level_id',)

years_serializer = serializers.ListField(child=serializers.IntegerField(),
                                         help_text='years to incude in the template',)

prognosis_id_serializer = serializers.PrimaryKeyRelatedField(
    queryset=Prognosis.objects.all(),
    required=False,
    help_text='prognosis_id',)


class PopulationTemplateSerializer(serializers.Serializer):
    """Serializer for uploading Population and Prognosis"""
    excel_file = serializers.FileField()
    drop_constraints = serializers.BooleanField(default=False)
    logger = logging.getLogger('population')
    scope = ProcessScope.POPULATION

    def create_template(self,
                        area_level_id: int,
                        years: List[int],
                        prognosis_id: int,
                        ) -> bytes:
        """Create a template file for population/prognosis"""
        area_level = AreaLevel.objects.get(pk=area_level_id)
        if prognosis_id is None:
            excel_filename = 'Einwohnerrealdaten.xlsx'
            title = 'Reale Einwohnerdaten'
        else:
            prognosis = Prognosis.objects.get(pk=prognosis_id)
            excel_filename = 'Prognosedaten.xlsx'
            title = f'Prognosevariante {prognosis.name}'

        dv_pos_float = DataValidation(type="decimal",
                                      operator="greaterThanOrEqual",
                                      formula1=0,
                                      allow_blank=True)
        dv_pos_float.error = 'Nur Zahlen >= 0 erlaubt'
        dv_pos_float.title = 'Ungültige positive Zahlen-Werte'


        # create Row-Multiindex for Areas with key and label
        areas = Area.annotated_qs(area_level)
        key_attr = get_area_level_key(self.logger, area_level)
        label_attr = get_area_level_label(self.logger, area_level)

        keys = ['id', key_attr, label_attr]
        idx_areas = pd.MultiIndex.from_frame(pd.DataFrame(areas.values(*keys), columns=keys))

        # create column-Multiindex with genders and agegroups
        genders = pd.DataFrame(Gender.objects.values_list('id', 'name'),
                               columns=['id', 'Geschlecht'])\
            .set_index('id')
        age_groups = pd.DataFrame([[ag.id, ag.name] for ag in AgeGroup.objects.all()],
                                  columns=['id', 'Altersgruppe']).set_index('id')

        idx_cols = pd.MultiIndex.from_product([genders.index, age_groups.index],
                                              names=['gender_id', 'age_group_id'])
        df_cols = idx_cols.to_frame(index=False)\
            .merge(age_groups, left_on='age_group_id', right_index=True)\
            .merge(genders, left_on='gender_id', right_index=True)
        idx_cols = pd.MultiIndex.from_frame(df_cols[['gender_id', 'Geschlecht',
                                                 'age_group_id', 'Altersgruppe']])

        # create an empty dataframe
        df_areas = pd.DataFrame(index=idx_areas,
                                columns=idx_cols)
        if not isinstance(years, list):
            years = [int(y) for y in years.split(',')]


        fn = os.path.join(settings.MEDIA_ROOT, excel_filename)
        with pd.ExcelWriter(fn, engine='openpyxl') as writer:
            for year in years:
                columns = ['area_id', 'gender_id', 'age_group_id', 'value']
                try:
                    population = Population.objects.get(
                        prognosis_id=prognosis_id, year__year=year,
                    )
                    rows = PopulationEntry.objects.filter(
                        population=population).values(*columns)
                except Population.DoesNotExist:
                    rows = []
                df_values = pd.DataFrame(
                    rows, columns=columns).rename(
                        columns={'area_id': 'id',}).set_index('id')

                if len(df_values):
                    df_values = df_values.pivot_table(
                        values='value', index=df_values.index,
                        columns=['gender_id', 'age_group_id'])
                    df_areas.loc[df_values.index, :] = df_values.values
                else:
                    df_areas.loc[:, :] = pd.NA


                df_areas_without_id = df_areas.droplevel('id')
                df_areas_without_id.to_excel(writer,
                                             sheet_name=str(year),
                                             startrow=1,
                                             freeze_panes=(6, 2))

                ws: Worksheet = writer.sheets.get(str(year))
                ws['A1'] = 'Jahr'
                ws['B1'] = year
                ws['C1'] = title
                ws.merge_cells(start_row=1, end_row=1,
                               start_column=3, end_column=df_areas.shape[1] + 2)

                row12 = ws[1:6]
                for row in row12:
                    for cell in row:
                        cell.alignment = styles.Alignment(horizontal='center',
                                                          wrap_text=True)

                from_row = 7
                to_row = from_row + df_areas.shape[0] - 1

                from_col_no = 3
                from_col = get_column_letter(from_col_no)
                to_col_no = from_col_no + df_areas.shape[1] - 2
                to_col = get_column_letter(to_col_no)

                dv_pos_float.add(f'{from_col}{from_row}:{to_col}{to_row}')
                ws.add_data_validation(dv_pos_float)

                ws.row_dimensions[2] = RowDimension(ws, index=2, hidden=True)
                ws.row_dimensions[4] = RowDimension(ws, index=4, hidden=True)

                ws.column_dimensions['A'] = ColumnDimension(ws, index='B',
                                                            width=20)
                ws.column_dimensions['B'] = ColumnDimension(ws, index='C',
                                                            width=30)
                for col_no in range(3, 3 + df_areas.shape[1]):
                    col = get_column_letter(col_no)
                    ws.column_dimensions[col] = ColumnDimension(ws, index=col,
                                                                width=12)

        content = open(fn, 'rb').read()
        return content


def get_area_level_label(logger, area_level: AreaLevel) -> str:
    try:
        label_attr = AreaField.objects.get(area_level=area_level,
                                           is_label=True).name
    except AreaField.DoesNotExist:
        msg = f'kein Label-Feld für Gebietseinheit {area_level.name} definiert'
        logger.error(msg)
        raise BadRequest(msg)
    return label_attr


def get_area_level_key(logger, area_level: AreaLevel) -> str:
    try:
        key_attr = AreaField.objects.get(area_level=area_level,
                                         is_key=True).name
    except AreaField.DoesNotExist:
        msg = f'Template konnte nicht erstellt werden, weil kein Schlüsselfeld für die Gebietseinheit {area_level.name} definiert ist.'
        logger.error(msg)
        raise BadRequest(msg)
    return key_attr

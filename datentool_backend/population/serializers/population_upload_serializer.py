from io import StringIO
import os
from collections import OrderedDict
from typing import Tuple, List
import pandas as pd


from django.conf import settings
from django.db.models.fields import FloatField
from rest_framework import serializers

from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl import styles
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.dimensions import ColumnDimension, RowDimension
from openpyxl.worksheet.datavalidation import DataValidation

from datentool_backend.area.models import AreaLevel, Area, AreaField
from datentool_backend.demand.models import Gender, AgeGroup
from datentool_backend.population.models import Population, Prognosis, PopulationEntry



area_level_id_serializer = serializers.PrimaryKeyRelatedField(
    queryset=AreaLevel.objects.all(),
    help_text='area_level_id',)

years_serializer = serializers.ListField(child=serializers.IntegerField(),
                                         help_text='years to incude in the template',)


prognosis_id_serializer = serializers.PrimaryKeyRelatedField(
    queryset=Prognosis.objects.all(),
    help_text='prognosis_id',)


class PopulationTemplateSerializer(serializers.Serializer):
    """Serializer for uploading Population and Prognosis"""
    excel_file = serializers.FileField()
    drop_constraints = serializers.BooleanField(default=True)

    def create_template(self,
                        area_level_id: int,
                        years: List[int],
                        prognosis_id: int,
                        ) -> bytes:
        """Create a template file for population/prognosis"""

        area_level = AreaLevel.objects.get(pk=area_level_id)
        if prognosis_id is None:
            excel_filename = 'Einwohnerrealdaten.xlsx'
            title = 'Reale Einwohnerdaten'
        else:
            prognosis = Prognosis.objects.get(pk=prognosis_id)
            excel_filename = 'Prognosedaten.xlsx'
            title = f'Prognosevariante {prognosis.name}'

        dv_pos_float = DataValidation(type="decimal",
                                operator="greaterThanOrEqual",
                                formula1=0,
                                allow_blank=True)
        dv_pos_float.error = 'Nur Zahlen >= 0 erlaubt'
        dv_pos_float.title = 'Ungültige positive Zahlen-Werte'


        # create Row-Multiindex for Areas with key and label
        areas = Area.annotated_qs(area_level)
        key_attr = AreaField.objects.get(area_level=area_level, is_key=True).name
        label_attr = AreaField.objects.get(area_level=area_level, is_label=True).name

        keys = ['id', key_attr, label_attr]
        idx_areas = pd.MultiIndex.from_frame(pd.DataFrame(areas.values(*keys), columns=keys))

        # create column-Multiindex with genders and agegroups
        genders = pd.DataFrame(Gender.objects.values_list('id', 'name'),
                               columns=['id', 'Geschlecht'])\
            .set_index('id')
        age_groups = pd.DataFrame([[ag.id, ag.name] for ag in AgeGroup.objects.all()],
                                  columns=['id', 'Altersgruppe']).set_index('id')

        idx_cols = pd.MultiIndex.from_product([genders.index, age_groups.index],
                                              names=['gender_id', 'age_group_id'])
        df_cols = idx_cols.to_frame(index=False)\
            .merge(age_groups, left_on='age_group_id', right_index=True)\
            .merge(genders, left_on='gender_id', right_index=True)
        idx_cols = pd.MultiIndex.from_frame(df_cols[['gender_id', 'Geschlecht',
                                                 'age_group_id', 'Altersgruppe']])

        # create an empty dataframe
        df_areas = pd.DataFrame(index=idx_areas,
                                columns=idx_cols)


        fn = os.path.join(settings.MEDIA_ROOT, excel_filename)
        with pd.ExcelWriter(fn, engine='openpyxl') as writer:
            meta = writer.book.create_sheet('meta')
            meta['A1'] = 'arealevel'
            meta['B1'] = area_level_id
            meta['A2'] = 'years count'
            meta['B2'] = len(years)
            meta['A3'] = 'years'
            for i, year in enumerate(years):
                meta.cell(3, i+2, year)
            meta['A4'] = 'prognosis'
            meta['B4'] = prognosis_id

            for year in years:

                df_areas.to_excel(writer,
                                  sheet_name=str(year),
                                  startrow=1,
                                  freeze_panes=(6, 3))

                ws: Worksheet = writer.sheets.get(str(year))
                ws['B1'] = 'Jahr'
                ws['C1'] = year
                ws['D1'] = title
                ws.merge_cells(start_row=1, end_row=1,
                               start_column=4, end_column=df_areas.shape[1] + 3)

                row12 = ws[1:6]
                for row in row12:
                    for cell in row:
                        cell.alignment = styles.Alignment(horizontal='center', wrap_text=True)

                from_row = 7
                to_row = from_row + df_areas.shape[0] - 1

                from_col_no = 4
                from_col = get_column_letter(from_col_no)
                to_col_no = from_col_no + df_areas.shape[1] - 1
                to_col = get_column_letter(to_col_no)

                dv_pos_float.add(f'{from_col}{from_row}:{to_col}{to_row}')
                ws.add_data_validation(dv_pos_float)

                ws.row_dimensions[2] = RowDimension(ws, index=2, hidden=True)
                ws.row_dimensions[4] = RowDimension(ws, index=4, hidden=True)

                ws.column_dimensions['A'] = ColumnDimension(ws, index='A', hidden=True)
                ws.column_dimensions['B'] = ColumnDimension(ws, index='B', width=20)
                ws.column_dimensions['C'] = ColumnDimension(ws, index='C', width=30)
                for col_no in range(4, 4 + df_areas.shape[1]):
                    col = get_column_letter(col_no)
                    ws.column_dimensions[col] = ColumnDimension(ws, index=col, width=12)

        content = open(fn, 'rb').read()
        return content

    def read_excel_file(self, request) -> pd.DataFrame:
        """read excelfile and return a dataframe"""
        excel_file = request.FILES['excel_file']

        # check if the right Excel-File is uploaded
        df_meta = pd.read_excel(excel_file.file,
                           sheet_name='meta').set_index('key')

        df = pd.DataFrame()
        return df
